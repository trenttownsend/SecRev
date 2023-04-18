import csv
import os
import sys
import subprocess
import re
import openpyxl
import io
import argparse
from argparse import Action
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime

os.system('cls')
print("SecRev by TT. 04-2023")
print("")

timestamp = datetime.now().strftime("%Y%m%d")
skipped = []
share_data = []
groups_data = []
inactive_users_data = []
disabled_users_data = []
filtered_groups = []
groups_no_members = []

class ServerSharesAction(Action):
    def __call__(self, parser, namespace, values, option_string=None):
        if not hasattr(namespace, "shares_by_server"):
            setattr(namespace, "shares_by_server", [])
        server = values[0]
        shares = tuple(values[1:])
        namespace.shares_by_server.append((server, shares))

parser = argparse.ArgumentParser(description="Security Review Tool")
parser.add_argument("-saveto", default="C:\\Temp", help="Save directory (optional)")
parser.add_argument("-s", "--server", nargs="+", action=ServerSharesAction, metavar=("SERVER", "Share"), help="Specify server and share(s)")
parser.add_argument("-workgroup", "-local", action="store_true", help="Set workgroup/local mode (default: False) Not currently implemented.")

if len(sys.argv) == 1:
    parser.print_help()
    sys.exit(1)

args = parser.parse_args()

saveto = args.saveto
output_filename = os.path.join(saveto, f"combined-output-{timestamp}.xlsx")
print(f"Saving to {output_filename}")

shares_by_server = getattr(args, "shares_by_server", [])
workgroup_local = args.workgroup
print("Workgroup =", workgroup_local)

filter_users = []
try:
    with open('ignore-users.txt', 'r') as ignore_users_file:
        filter_users = [g.rstrip('\n') for g in ignore_users_file.readlines()]
except:
    print("""
    No ignore-users.txt file found.
          
    Continuing without filtering.
    ==================================================================================================
    """)
    pass

filter_groups = []
try:
    with open('ignore-groups.txt', 'r') as ignore_groups_file:
        filter_groups = [g.rstrip('\n') for g in ignore_groups_file.readlines()]
except:
    print("""
    No ignore-groups.txt file found.
          
    Continuing without filtering.
    ==================================================================================================
    """)
    pass
convert_users = []
try:
    with open('convert-users.txt', 'r') as convert_file:
        convert_users = [g.rstrip('\n') for g in convert_file.readlines()]
except:
    print("""
    No convert-users.txt file found.
    If you want to convert users, create a file called convert-users.txt and add the following format:
    <user to convert>, <new user>
    
    Continuing without conversion.
    ==================================================================================================
    """)
    pass

def simplify_permissions(permissions):
    simplified_permissions = set()
    ntfs_permissions = {
        "2032127": "FullControl",
        "1179817": "ReadAndExecute",
        "1179808": "Read",
        "1180080": "Write",
        "1180095": "Append",
        "2032128": "Modify"
    }

    for permission in permissions:
        if permission in ntfs_permissions:
            simplified_permissions.add(ntfs_permissions[permission])
        elif permission.isdigit():
            binary = bin(int(permission))
            binary_str = str(binary)[2:].rjust(32, '0')
            perm_str = binary_str[-3:]
            if perm_str == "001":
                simplified_permissions.add("ReadAndExecute")
            elif perm_str == "010":
                simplified_permissions.add("Read")
            elif perm_str == "011":
                simplified_permissions.add("Read+Execute")
            elif perm_str == "100":
                simplified_permissions.add("Write")
            elif perm_str == "101":
                simplified_permissions.add("Write+Execute")
            elif perm_str == "110":
                simplified_permissions.add("Write+Read")
            elif perm_str == "111":
                simplified_permissions.add("FullControl")
            else:
                continue
        else:
            if permission == "FullControl":
                simplified_permissions.add("FullControl")
            elif permission == "Read, Synchronize":
                simplified_permissions.add("Read")
            elif permission == "ReadAndExecute, Synchronize":
                simplified_permissions.add("Read+Execute")
            elif permission == "Read":
                simplified_permissions.add("Read")
            elif "Modify" in permission or "Write" in permission or "Delete" in permission:
                simplified_permissions.add("Read+Write")

    return ", ".join(sorted(simplified_permissions))

def convert(csv_data, disabledusers = [], allowedusers = [], usersgroups = {}):
    columns = set()
    user_map = {}

    csvfile = io.StringIO(csv_data)
    reader = list(csv.reader(csvfile, delimiter=","))

    #print("UG MAP:", user_group_map)

    for row in reader[2:]:
        access_type = row[2]
        if access_type.lower().strip() == "deny":
            continue
        col = row[0]
        user = row[3]

        #allowedusers_lower = {account.lower() for account in allowedusers}
        #if "everyone" not in allowedusers_lower and "domain users" not in allowedusers_lower and usersgroups != {}:
        #    if user.lower() not in allowedusers_lower:
        #        user_groups = {group for group, users in user_group_map.items() if user.lower() in users}
        #        if not user_groups.intersection(allowedusers_lower):
        #            continue

        cont = True
        if allowedusers != []:
            allowed_accounts = {account.lower() for account in allowedusers}
            #print("ALLOWED ACCOUNTS:", allowed_accounts)
            if "everyone" not in allowed_accounts and "domain users" not in allowed_accounts:
                cont = True
            else:
                cont = False
        if usersgroups != {} and cont == True:
            user_group_members = {usersgroups.get(group.lower(), []) for group in allowed_accounts}
            user_group_members_flat = {member for group_members in user_group_members for member in group_members}
            #print("FLAT:", user_group_members_flat)
            if not (user.lower() in allowed_accounts or user.lower() in user_group_members_flat):
                continue

        if user.lower in disabledusers:
            continue
        elif user.startswith("S-1-5-21"):
            continue

        permission = row[1]
        columns.add(col)
        for users in convert_users:
            if user.lower() in users.lower():
                user = users.split(', ')[1]
        if user not in user_map:
            user_map[user] = {}
        if col not in user_map[user]:
            user_map[user][col] = set()
        user_map[user][col].add(permission)

    shareName = ""
    for c in columns:
        if shareName == "":
            shareName = re.match(r'^\\\\(.*?)\\(.*?)\\', c).group(0)

    columns = sorted(columns, key=lambda s: s.lower())
    columns2 = [re.sub(r'^\\\\(.*?)\\(.*?)\\', '', c) for c in columns]
    columns2 = sorted(columns2, key=lambda s: s.lower())
    column_users = [{} for _ in columns]

    for user in user_map:
        if user in filter_users:
            continue
        for col in user_map[user]:
            idx = columns.index(col)
            column_users[idx][user] = simplify_permissions(user_map[user][col])

    same_permissions = all(user_dict == column_users[0] for user_dict in column_users[1:])

    if same_permissions:
        columns2 = ['All Subfolders']
        column_users = [column_users[0]]

    result = []
    result.append([shareName, *columns2])

    max_rows = max(len(users) for users in column_users)
    for row_idx in range(max_rows):
        row = ['']
        for users in column_users:
            user_list = sorted(users.keys())
            if row_idx < len(user_list):
                user = user_list[row_idx]
                permissions = users[user]
                row.append(f"{user} ({permissions})")
            else:
                row.append('')
        result.append(row)

    return result

def write_inactive_users_to_excel(inactive_users, output_filename):
    try:
        wb = load_workbook(output_filename)
    except:
        wb = openpyxl.Workbook()
        pass
    if 'Inactive Users' in wb.sheetnames:
        ws = wb['Inactive Users']
    else:
        ws = wb.create_sheet('Inactive Users')
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    header_fill = PatternFill(start_color='F79646', end_color='F79646', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row_idx, row in enumerate(inactive_users, start=2):  # Skip the first two rows and start writing from the 2nd row
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = Font(name='Calibri', size=10)

            if row_idx == 2:
                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.border = border
            
            if col_idx == 3 and cell.value == '':
                cell.value = 'N/A'

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(output_filename)
    return True

def write_groups_to_excel(groups, output_filename, disabledusers = []):
    try:
        wb = load_workbook(output_filename)
    except:
        wb = openpyxl.Workbook()
        pass
    if 'Groups' in wb.sheetnames:
        ws = wb['Groups']
    else:
        ws = wb.create_sheet('Groups')
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    header_fill = PatternFill(start_color='F79646', end_color='F79646', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for group_name in groups[1:]:
        if group_name[0] in filter_groups:
            try:
                groups.remove(group_name)
                filtered_groups.append(group_name)
                #print("Group removed: " + group_name[0] + " - " + group_name[1])
            except:
                pass
        elif not group_name[1] or group_name[1] == '':
            try:
                groups.remove(group_name)
                filtered_groups.append(group_name)
                #print("Group removed: " + group_name[0] " - No members")
            except:
                pass
    groups = sorted(groups[1:], key=lambda x: x[0])
    groups = sorted(groups, key=lambda x: x[1])

    # Write the group names as headers in row 2
    cell = ws.cell(row=2, column=1, value="Groups:")
    cell.font = Font(name='Calibri', size=10, underline='single', bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.border = border
    
    for group in groups:
        for user in disabledusers:
            memblist = list(group[1].split(', '))
            for members in memblist:
                if members.lower() == user and len(memblist) == 1:
                    try:
                        groups.remove(group)
                    except:
                        print('Error removing group: ' + group[0])
                        pass
                elif members.lower() == user and len(memblist) > 1:
                    group[1] = group[1].replace((user + ', '), '')
                    
        if not group[1] or group[1] == '':
            try:
                groups.remove(group)
                groups_no_members.append(group)
            except:
                print('Error removing group: ' + group[0])
                pass
    
    groups.sort(key=lambda x: x[0])

    for col_idx, group_name in enumerate(groups, start=2):
        cell = ws.cell(row=2, column=col_idx, value=group_name[0])
        cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.border = border

        # Add group members under their respective groups
        userlist = []

        for user in group_name[1].split(', '):
            if user.lower() not in disabledusers:
                #print(user)
                userlist.append(user)
            else:
                continue
        userlist.sort()

        for user in userlist:
            row_idx = 3
            while ws.cell(row=row_idx,column=col_idx).value:  # Skip non-blank rows
                row_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=user).font = Font(name='Calibri', size=10)

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(output_filename)
    return True

def write_to_excel(data, output_filename):
    try:
        wb = load_workbook(output_filename)  # Load the template file
    except:
        wb = openpyxl.Workbook()  # Create a new workbook if the template file is not found
        pass
    if 'Shares' in wb.sheetnames:
        ws = wb['Shares']  # Use the 'Shares' sheet if it exists
    else:
        ws = wb.create_sheet('Shares')
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    header_fill = PatternFill(start_color='F79646', end_color='F79646', fill_type='solid')  # Accent 6 fill color
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row_idx, row in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = Font(name='Calibri', size=10)

            # Apply formatting to the header rows
            if (row_idx == 1 or (col_idx == 1 and cell_value and '\\' in cell_value)):
                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')  # Bold, White font, Calibri, size 10
                cell.fill = header_fill
                cell.border = border
            elif cell_value.lower() == 'all subfolders':
                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                cell.fill = header_fill
                cell.border = border
            elif ws.cell(row=row_idx, column=1).value and '\\' in ws.cell(row=row_idx, column=1).value:
                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')  # Bold, White font, Calibri, size 10
                cell.fill = header_fill
                cell.border = border

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(output_filename)
    return True

# process CSV data
def read_csv(csv_data, name):
    #print("CSV DATA:", csv_data)
    csvfile = io.StringIO(csv_data)
    reader = list(csv.reader(csvfile, delimiter=","))
    data = [row for row in reader]
    #print("PROCESSED DATA:", data)
    if data:
        return data
    else:
        print('No CSV data found in', name)

# call share-permissions.ps1
def processShares(ps_cmd, srv, shr, UGs):
    try:
        allowed_users = []
        process_Shares = subprocess.Popen(ps_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        stdout, stderr = process_Shares.communicate()
        sharesLines = stdout
        #print("ShareLines:", sharesLines)
        share_access_start = sharesLines.find("###SHARE_ACCESS_START###") + len("###SHARE_ACCESS_START###")
        share_access_end = sharesLines.find("###SHARE_ACCESS_END###")
        share_access_csv = sharesLines[share_access_start:share_access_end].strip()
        #print("share_access_csv:", share_access_csv)
        try:
            share_access_rows = read_csv(share_access_csv, "access rights for \\\\{0}\{1}".format(srv, shr))
            #print("share_access_rows:", share_access_rows)
            if len(share_access_rows) <= 1:  # If only the header row is present
                allowed_users.append("everyone")
            else:
                allowed_users = [row[2] for row in share_access_rows if row[0].strip().lower() == 'allow']
        except Exception as e:
            print("Error processing access data for share \\\\{0}\{1}: {2}".format(srv, shr, e))
            allowed_users.append("everyone")
        #print("Allowed:", allowed_users)
        share_data_start = sharesLines.find("###SHARE_DATA_START###") + len("###SHARE_DATA_START###")
        share_data_end = sharesLines.find("###SHARE_DATA_END###")
        share_data_csv = sharesLines[share_data_start:share_data_end].strip()
        if not share_data_csv or share_data_csv.count('\n') == 1:
            print("No subfolders found for share \\\\{0}\{1}".format(srv, shr))
            skipped.append("Share \\\\{0}\{1}".format(srv, shr))
            return
    except:
        print("Error executing share-permissions.ps1 (PowerShell)")
        print("Failed execution:", ps_cmd + "")

    try:
        share_data.extend([[]])  # Add two blank lines between shares
        converted_data = convert(share_data_csv, disabled_users_data, allowed_users, UGs)
        share_data.extend(converted_data)
        share_data.extend([[]])  # Add two blank lines between shares
        print("Appended share \\\\{0}\{1}".format(srv, shr))
    except Exception as e:
        print("Error appending share \\\\{0}\{1}: {2}".format(srv, shr, e))
        skipped.append("Share \\\\{0}\{1}".format(srv, shr))

script_dir = ''

if getattr(sys, 'frozen', False):
    # Running in PyInstaller bundle
    script_dir = sys._MEIPASS
else:
    # Running in normal Python environment
    script_dir = '.'

users_groups_script = os.path.join(script_dir, 'users-groups.ps1')
share_permissions_script = os.path.join(script_dir, 'share-permissions.ps1')

# users-groups.ps1 - Get all users and groups
ps_cmd_ugs = f"powershell.exe -NoLogo -ExecutionPolicy ByPass -File {users_groups_script}"

# Execute the PowerShell command and wait for it to complete
print("")
print("Executing users-groups.ps1 (PowerShell)")
try:
    #processUsers = subprocess.Popen(ps_cmd_ugs, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    #outputUGs, errorsUGs = processUsers.communicate()
    #UGsLines = outputUGs.decode('utf-8')
    processUsers = subprocess.Popen(ps_cmd_ugs, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    stdout, stderr = processUsers.communicate()
    UGsLines = stdout
    #print(UGsLines)
except:
    print("Error executing users-groups.ps1 (PowerShell)")
    print("Failed execution: " + ps_cmd_ugs + "")

# Parsing the output using the find method
try:
    inactive_start = UGsLines.find("###INACTIVE_USERS_START###") + len("###INACTIVE_USERS_START###")
    inactive_end = UGsLines.find("###INACTIVE_USERS_END###")
    inactive_csv = UGsLines[inactive_start:inactive_end].strip()
    #print(inactive_csv)
    if inactive_csv:
        inactive_users_data.extend(read_csv(inactive_csv, "inactive users"))
        print("Extending inactive users")
    else:
        print("Error retrieving inactive users")
        skipped.append("inactive users")
except:
    print("Error retrieving inactive users")
    skipped.append("inactive users")

try:
    disabled_start = UGsLines.find("###DISABLED_USERS_START###") + len("###DISABLED_USERS_START###")
    disabled_end = UGsLines.find("###DISABLED_USERS_END###")
    disabled_csv = UGsLines[disabled_start:disabled_end].strip()
    #print(disabled_csv)
    if disabled_csv:
        disabled_users_data.extend(read_csv(disabled_csv, "disabled users"))
        disabled_users = []
        if disabled_users_data[1]:
            for user in disabled_users_data[1:]:
                disabled_users.append(user[0].lower())
        disabled_users_data = disabled_users
        #print(disabled_users_data)
        print("Extending disabled users")
    else:
        print("Error retrieving disabled users")
        skipped.append("disabled users")
except:
    print("Error retrieving disabled users")
    skipped.append("disabled users")

user_group_map = {}

try:
    group_members_start = UGsLines.find("###GROUP_MEMBERS_START###") + len("###GROUP_MEMBERS_START###")
    group_members_end = UGsLines.find("###GROUP_MEMBERS_END###")
    group_members_csv = UGsLines[group_members_start:group_members_end].strip()

    if group_members_csv:
        groups_data.extend(read_csv(group_members_csv, "groups"))
        for group, users_str in groups_data:
            users_list = users_str.split(', ')
            user_group_map[group] = users_list
        print("Extending groups & memberships")
    else:
        print("Error retrieving groups & memberships")
        skipped.append("groups & memberships")
except:
    print("Error retrieving groups & memberships")
    skipped.append("groups & memberships")


# share-permissions.ps1 - Get all share permissions
print("")
print("Executing share-permissions.ps1 (PowerShell)")
for server in shares_by_server:
    # shares_by_server = tuple of (server, list of shares)
    # share = (server, list of shares) 
    for share in server[1]:
        sharecommand = f'''powershell.exe -NoLogo -ExecutionPolicy ByPass -File {share_permissions_script} -Fileserver "{server[0]}" -Share "{share}"'''
        processShares(sharecommand, server[0], share, user_group_map)


# Create the destination directory if it doesn't exist
if not os.path.exists(saveto):
    os.makedirs(saveto)

if os.path.exists(os.path.join(saveto, output_filename)):
    try:
        os.remove(os.path.join(saveto, output_filename))
    except:
        print("Error removing existing output file")
        print("Please close the file and try again")
        sys.exit(1)

write_inactive_users_to_excel(inactive_users_data, output_filename)
write_groups_to_excel(groups_data, output_filename, disabled_users_data)
write_to_excel(share_data, output_filename)

if skipped:
    print("")
    print("Some or all of the following data may not have been collected.")
    print("")
for skip in skipped:
    print("Skipped", skip)
print("")
### for grp in filtered_groups:
###     if grp[1]:
###         print("Group '" + grp[0] + "' with", len(grp[1]), "members filtered from output")
###     elif not grp[1] or grp[1] == '':
###         print("Group '" + grp[0] + "'with no members filtered from output")
### for grp in groups_no_members:
###     print("Excluded group '" + grp + "' from output - no members")
print("")
print('Done!')
sys.exit(0)
